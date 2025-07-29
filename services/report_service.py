"""Report service for filtering and analyzing Fieldwire project data."""

import pandas as pd
from typing import Dict, List, Optional
import openpyxl
from enum import Enum, auto

class FilterType(Enum):
    """Types of filters available."""
    TASKS = auto()
    CHECK_ITEMS = auto()

class TaskFilterType(Enum):
    """Types of task filters available."""
    TEAM = auto()
    STATUS = auto()
    ATTRIBUTE = auto()

class CheckItemFilterType(Enum):
    """Types of check item filters available."""
    STATE = auto()
    NAME = auto()

class TextMatchType(Enum):
    """Types of text matching available for names."""
    CONTAINS = auto()
    STARTS_WITH = auto()
    ENDS_WITH = auto()
    EXACTLY_MATCHES = auto()

class FilterState:
    """Maintains the current state of filters being applied."""
    
    def __init__(self):
        self.task_filters = []
        self.check_item_filters = []

class ReportService:
    """Service for filtering and reporting on Fieldwire project data."""
    
    # Check item states and their display values
    CHECK_ITEM_STATES = {
        'empty': 'pending',
        'yes': 'complete',
        'no': 'missing',
        'not_applicable': 'greyed'
    }
    
    # Colors for check item states in Excel
    CHECK_ITEM_COLORS = {
        'empty': '#FFFFFF',        # White for pending
        'yes': '#90EE90',         # Light green for complete
        'no': '#FFB6C6',          # Light red for missing
        'not_applicable': '#D3D3D3'  # Light grey for not applicable
    }
    
    def __init__(self, project_service, task_service, attribute_service, 
                 status_service, team_service, tag_service):
        """Initialize the report service.
        
        Args:
            project_service: ProjectService instance
            task_service: TaskService instance
            attribute_service: AttributeService instance
            status_service: StatusService instance
            team_service: TeamService instance
            tag_service: TagService instance
        """
        self.project_service = project_service
        self.task_service = task_service
        self.attribute_service = attribute_service
        self.status_service = status_service
        self.team_service = team_service
        self.tag_service = tag_service
        
        self.dataframes: Dict[str, pd.DataFrame] = {}
        self.filter_state = FilterState()
        
    def initialize_project_data(self, project_name: str) -> None:
        """Initialize all DataFrames for a given project."""
        # Get project ID from name
        project_id = self.project_service.get_project_id_from_name_or_id(project_name)
        if not project_id:
            raise ValueError(f"Project '{project_name}' not found")
            
        print("\nFetching project data...")
        
        try:
            # Fetch all required data
            tasks = self.task_service.get_all_tasks_in_project(project_id, filter_option='active')
            teams = self.team_service.get_all_teams_in_project(project_id)
            statuses = self.status_service.get_statuses_for_project_id(project_id)
            task_check_items = self.attribute_service.get_all_task_check_items_in_project(project_id)
            task_type_attributes = self.attribute_service.get_all_task_type_attributes_in_project(project_id)
            task_attributes = self.attribute_service.get_all_task_attributes_in_project(project_id)
            
            # Convert to DataFrames
            if not all([tasks, teams, statuses, task_check_items, task_type_attributes, task_attributes]):
                raise ValueError("Failed to fetch required data")
                
            # Convert task_check_items to DataFrame if it's not already
            if isinstance(task_check_items, list):
                task_check_items_df = pd.DataFrame(task_check_items)
            else:
                task_check_items_df = pd.DataFrame([task_check_items])
                
            self.dataframes = {
                'tasks': pd.DataFrame(tasks)[['id', 'team_id', 'status_id', 'name']],
                'teams': pd.DataFrame(teams)[['id', 'name']],
                'statuses': pd.DataFrame(statuses)[['id', 'name']],
                'task_check_items': task_check_items_df[['id', 'task_id', 'state', 'name']],
                'task_type_attributes': pd.DataFrame(task_type_attributes)[['id', 'name']],
                'task_attributes': pd.DataFrame(task_attributes)[['id', 'task_id', 'task_type_attribute_id', 'text_value']]
            }
            
            # Convert all values to strings
            for name, df in self.dataframes.items():
                self.dataframes[name] = df.astype(str).replace('nan', '')
                
            print("Data fetched and processed successfully.")
            
        except Exception as e:
            raise ValueError(f"Failed to initialize project data: {str(e)}")
            
    def start_filtering(self) -> None:
        """Start the filtering process by asking the user what to filter."""
        print("\nWhat would you like to do?")
        print("1. Filter Tasks")
        print("2. Filter Check Items")
        print("3. Export All Raw Data to Excel")
        
        choice = input("Enter your choice (1-3): ").strip()
        
        if choice == "1":
            self._handle_task_filtering()
        elif choice == "2":
            self._handle_check_item_filtering()
        elif choice == "3":
            filename = input("\nEnter filename for Excel export (e.g., raw_data.xlsx): ").strip()
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            self.export_raw_data(filename)
        else:
            print("Invalid choice. Please try again.")
            self.start_filtering()
            
    def _handle_task_filtering(self) -> None:
        """Handle the task filtering process."""
        while True:
            print("\nChoose task filter type:")
            print("1. Team")
            print("2. Status")
            print("3. Task Attribute")
            print("4. Done with task filters")
            
            choice = input("Enter your choice (1-4): ").strip()
            
            if choice == "4":
                break
                
            if choice == "1":
                self._add_team_filter()
            elif choice == "2":
                self._add_status_filter()
            elif choice == "3":
                self._add_task_attribute_filter()
            else:
                print("Invalid choice. Please try again.")
                
    def _handle_check_item_filtering(self) -> None:
        """Handle the check item filtering process."""
        while True:
            print("\nChoose check item filter type:")
            print("1. State")
            print("2. Name")
            print("3. Done with check item filters")
            
            choice = input("Enter your choice (1-3): ").strip()
            
            if choice == "3":
                # Move on to task filtering
                print("\nNow let's filter which tasks to look for these check items in.")
                self._handle_task_filtering()
                break
                
            if choice == "1":
                self._add_check_item_state_filter()
            elif choice == "2":
                self._add_check_item_name_filter()
            else:
                print("Invalid choice. Please try again.")
                
    def _add_team_filter(self) -> None:
        """Add a team filter."""
        teams = self.dataframes['teams']['name'].unique()
        print("\nAvailable teams:")
        for i, team in enumerate(teams, 1):
            print(f"{i}. {team}")
            
        try:
            choice = int(input("Enter team number: "))
            if 1 <= choice <= len(teams):
                selected_team = teams[choice - 1]
                self.filter_state.task_filters.append(
                    ('team', selected_team)
                )
                print(f"Added filter for team: {selected_team}")
            else:
                print("Invalid choice.")
        except ValueError:
            print("Please enter a valid number.")
            
    def _add_status_filter(self) -> None:
        """Add a status filter."""
        statuses = self.dataframes['statuses']['name'].unique()
        print("\nAvailable statuses:")
        for i, status in enumerate(statuses, 1):
            print(f"{i}. {status}")
            
        try:
            choice = int(input("Enter status number: "))
            if 1 <= choice <= len(statuses):
                selected_status = statuses[choice - 1]
                self.filter_state.task_filters.append(
                    ('status', selected_status)
                )
                print(f"Added filter for status: {selected_status}")
            else:
                print("Invalid choice.")
        except ValueError:
            print("Please enter a valid number.")
            
    def _add_task_attribute_filter(self) -> None:
        """Add a task attribute filter."""
        if 'task_type_attributes' not in self.dataframes or 'task_attributes' not in self.dataframes:
            print("\nNo task attributes found in the project.")
            return
            
        # Get available attribute types with their names
        attribute_types = self.dataframes['task_type_attributes']
        
        print("\nAvailable attribute types:")
        for i, (attr_id, attr_name) in enumerate(
            zip(attribute_types['id'], attribute_types['name']), 1
        ):
            print(f"{i}. {attr_name}")
            
        try:
            type_choice = int(input("Enter attribute type number: "))
            if 1 <= type_choice <= len(attribute_types):
                # Get the selected attribute type
                selected_type_id = attribute_types.iloc[type_choice - 1]['id']
                selected_type_name = attribute_types.iloc[type_choice - 1]['name']
                
                # Get unique values for this attribute type
                values = self.dataframes['task_attributes'][
                    self.dataframes['task_attributes']['task_type_attribute_id'] == selected_type_id
                ]['text_value'].unique()
                
                print(f"\nAvailable values for {selected_type_name}:")
                for i, value in enumerate(values, 1):
                    print(f"{i}. {value}")
                    
                value_choice = int(input("Enter value number: "))
                if 1 <= value_choice <= len(values):
                    selected_value = values[value_choice - 1]
                    self.filter_state.task_filters.append(
                        ('attribute', (selected_type_id, selected_value))
                    )
                    print(f"Added filter for {selected_type_name} = {selected_value}")
                else:
                    print("Invalid value choice.")
            else:
                print("Invalid type choice.")
        except ValueError:
            print("Please enter a valid number.")
            
    def _add_check_item_state_filter(self) -> None:
        """Add a check item state filter."""
        states = list(self.CHECK_ITEM_STATES.items())
        print("\nAvailable states:")
        for i, (state, display) in enumerate(states, 1):
            print(f"{i}. {display} ({state})")
            
        try:
            choice = int(input("Enter state number: "))
            if 1 <= choice <= len(states):
                selected_state = states[choice - 1][0]
                self.filter_state.check_item_filters.append(
                    ('state', selected_state)
                )
                print(f"Added filter for state: {self.CHECK_ITEM_STATES[selected_state]}")
            else:
                print("Invalid choice.")
        except ValueError:
            print("Please enter a valid number.")
            
    def _add_check_item_name_filter(self) -> None:
        """Add a check item name filter."""
        print("\nChoose text match type:")
        print("1. Contains")
        print("2. Starts with")
        print("3. Ends with")
        print("4. Exactly matches")
        
        match_type = input("Enter choice (1-4): ").strip()
        if match_type not in ["1", "2", "3", "4"]:
            print("Invalid choice.")
            return
            
        search_text = input("Enter search text: ").strip()
        if not search_text:
            print("Search text cannot be empty.")
            return
            
        match_types = {
            "1": "contains",
            "2": "starts_with",
            "3": "ends_with",
            "4": "exactly_matches"
        }
        
        self.filter_state.check_item_filters.append(
            ('name', (match_types[match_type], search_text))
        )
        print(f"Added name filter: {match_types[match_type]} '{search_text}'")
            
    def apply_filters(self) -> pd.DataFrame:
        """Apply all filters and return the resulting DataFrame."""
        # Start with all tasks
        filtered_tasks = self.dataframes['tasks'].copy()
        
        # Apply task filters
        for filter_type, filter_value in self.filter_state.task_filters:
            if filter_type == 'team':
                team_id = self.dataframes['teams'][
                    self.dataframes['teams']['name'] == filter_value
                ]['id'].iloc[0]
                filtered_tasks = filtered_tasks[filtered_tasks['team_id'] == team_id]
            elif filter_type == 'status':
                status_id = self.dataframes['statuses'][
                    self.dataframes['statuses']['name'] == filter_value
                ]['id'].iloc[0]
                filtered_tasks = filtered_tasks[filtered_tasks['status_id'] == status_id]
            elif filter_type == 'attribute':
                attr_type_id, attr_value = filter_value
                # Get task IDs that match the attribute filter
                matching_task_ids = self.dataframes['task_attributes'][
                    (self.dataframes['task_attributes']['task_type_attribute_id'] == attr_type_id) &
                    (self.dataframes['task_attributes']['text_value'] == attr_value)
                ]['task_id'].unique()
                filtered_tasks = filtered_tasks[filtered_tasks['id'].isin(matching_task_ids)]
        
        # Get check items for filtered tasks
        check_items = self.dataframes['task_check_items'][
            self.dataframes['task_check_items']['task_id'].isin(filtered_tasks['id'])
        ]
        
        # Apply check item filters
        for filter_type, filter_value in self.filter_state.check_item_filters:
            if filter_type == 'state':
                check_items = check_items[check_items['state'] == filter_value]
            elif filter_type == 'name':
                match_type, search_text = filter_value
                if match_type == 'contains':
                    check_items = check_items[check_items['name'].str.contains(search_text, na=False)]
                elif match_type == 'starts_with':
                    check_items = check_items[check_items['name'].str.startswith(search_text, na=False)]
                elif match_type == 'ends_with':
                    check_items = check_items[check_items['name'].str.endswith(search_text, na=False)]
                elif match_type == 'exactly_matches':
                    check_items = check_items[check_items['name'] == search_text]
        
        # Create final result
        result = pd.merge(
            filtered_tasks[['id', 'name']],
            check_items[['task_id', 'name', 'state']],
            left_on='id',
            right_on='task_id',
            suffixes=('_task', '_check_item')
        )
        
        # Rename columns
        result = result.rename(columns={
            'name_task': 'Task Name',
            'name_check_item': 'Check Item',
            'state': 'State'
        })
        
        # Drop ID columns and sort
        result = result.drop(columns=['id', 'task_id'])
        result = result.sort_values('Task Name')
        
        return result
            
    def export_to_excel(self, filename: str, df: pd.DataFrame) -> None:
        """Export DataFrame to Excel with formatting.
        
        Args:
            filename: Name of the Excel file to create
            df: DataFrame to export
            
        Raises:
            ValueError: If export fails
        """
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Export to Excel
                df.to_excel(writer, sheet_name='Report', index=False)
                
                # Get the worksheet
                worksheet = writer.sheets['Report']
                
                # Apply color coding for check item states if State column exists
                if 'State' in df.columns:
                    state_col_idx = df.columns.get_loc('State') + 1
                    for row_idx, value in enumerate(df['State'], start=2):
                        cell = worksheet.cell(row=row_idx, column=state_col_idx)
                        if value in self.CHECK_ITEM_COLORS:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=self.CHECK_ITEM_COLORS[value].replace('#', ''),
                                end_color=self.CHECK_ITEM_COLORS[value].replace('#', ''),
                                fill_type='solid'
                            )
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
                
                # Add alternating row colors
                prev_task = None
                row_color = 'F0F0F0'
                for row_idx, task in enumerate(df['Opening' if 'Opening' in df.columns else 'Task Name'], start=2):
                    if task != prev_task:
                        row_color = 'F0F0F0' if row_color == 'FFFFFF' else 'FFFFFF'
                    prev_task = task
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        if 'State' in df.columns and col_idx == df.columns.get_loc('State') + 1:
                            continue  # Skip state column as it has its own coloring
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=row_color,
                            end_color=row_color,
                            fill_type='solid'
                        )
                
                print(f"\nData exported successfully to {filename}")
                
        except Exception as e:
            raise ValueError(f"Failed to export to Excel: {str(e)}")

    def export_raw_data(self, filename: str) -> None:
        """Export all raw data to separate sheets in an Excel file.
        
        Args:
            filename: Name of the Excel file to create
            
        Raises:
            ValueError: If export fails
        """
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Export each DataFrame to its own sheet
                for name, df in self.dataframes.items():
                    # Create a copy to avoid modifying original
                    export_df = df.copy()
                    
                    # Convert all values to strings and replace empty values
                    export_df = export_df.astype(str).replace(['nan', ''], '')
                    
                    # Export to sheet
                    sheet_name = name.replace('_', ' ').title()
                    export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the worksheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
                    
                    # Add alternating row colors
                    for row_idx in range(2, len(export_df) + 2):  # Start from 2 to skip header
                        row_color = 'F0F0F0' if row_idx % 2 == 0 else 'FFFFFF'
                        for col_idx in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=row_color,
                                end_color=row_color,
                                fill_type='solid'
                            )
                
                print(f"\nAll raw data exported successfully to {filename}")
                print("The following sheets were created:")
                for name in self.dataframes.keys():
                    print(f"- {name.replace('_', ' ').title()}")
                
        except Exception as e:
            raise ValueError(f"Failed to export raw data to Excel: {str(e)}")

    def generate_fc_task_report(self, project_name: str, output_filename: str) -> None:
        """Generate a report of FC tasks with their attributes.
        
        Args:
            project_name: Name of the project to generate report for
            output_filename: Name of the Excel file to save the report to
            
        Raises:
            ValueError: If project not found or required data is missing
        """
        try:
            # Get project ID from name
            project_id = self.project_service.get_project_id_from_name_or_id(project_name)
            if not project_id:
                raise ValueError(f"Project '{project_name}' not found")
                
            print("\nFetching project data...")
            
            # Get all required data
            tasks = self.task_service.get_all_tasks_in_project(project_id, filter_option='active')
            task_type_attributes = self.attribute_service.get_all_task_type_attributes_in_project(project_id)
            task_attributes = self.attribute_service.get_all_task_attributes_in_project(project_id)
            
            if not all([tasks, task_type_attributes, task_attributes]):
                raise ValueError("Failed to fetch required data")
            
            # Convert to DataFrames
            tasks_df = pd.DataFrame(tasks)
            task_type_attributes_df = pd.DataFrame(task_type_attributes)
            task_attributes_df = pd.DataFrame(task_attributes)
            
            # Validate required task type attributes exist
            required_attributes = ["Strike Jamb", "Hinge Jamb", "Frame Header"]
            missing_attributes = [attr for attr in required_attributes 
                                if not any(task_type_attributes_df['name'] == attr)]
            
            if missing_attributes:
                raise ValueError(f"Missing required task type attributes: {', '.join(missing_attributes)}")
            
            # Filter for FC tasks
            fc_tasks = tasks_df[tasks_df['name'].str.contains('FC', na=False)]
            
            if len(fc_tasks) == 0:
                raise ValueError("No tasks found containing 'FC' in their name")
            
            print(f"\nFound {len(fc_tasks)} FC tasks. Processing...")
            
            # Create result DataFrame
            result_data = []
            
            # Process each FC task with progress bar
            from tqdm import tqdm
            for _, task in tqdm(fc_tasks.iterrows(), total=len(fc_tasks), desc="Processing tasks"):
                task_attrs = task_attributes_df[task_attributes_df['task_id'] == task['id']]
                
                # Get attribute values
                strike_jamb = ''
                hinge_jamb = ''
                frame_header = ''
                
                # Get attribute IDs
                strike_jamb_id = task_type_attributes_df[task_type_attributes_df['name'] == 'Strike Jamb']['id'].iloc[0] if not task_type_attributes_df[task_type_attributes_df['name'] == 'Strike Jamb'].empty else None
                hinge_jamb_id = task_type_attributes_df[task_type_attributes_df['name'] == 'Hinge Jamb']['id'].iloc[0] if not task_type_attributes_df[task_type_attributes_df['name'] == 'Hinge Jamb'].empty else None
                frame_header_id = task_type_attributes_df[task_type_attributes_df['name'] == 'Frame Header']['id'].iloc[0] if not task_type_attributes_df[task_type_attributes_df['name'] == 'Frame Header'].empty else None
                
                if not task_attrs.empty:
                    if strike_jamb_id:
                        strike_jamb_attr = task_attrs[task_attrs['task_type_attribute_id'] == strike_jamb_id]
                        if not strike_jamb_attr.empty:
                            strike_jamb = strike_jamb_attr['text_value'].iloc[0]
                            if isinstance(strike_jamb, dict):
                                strike_jamb = strike_jamb.get('value', '')
                    
                    if hinge_jamb_id:
                        hinge_jamb_attr = task_attrs[task_attrs['task_type_attribute_id'] == hinge_jamb_id]
                        if not hinge_jamb_attr.empty:
                            hinge_jamb = hinge_jamb_attr['text_value'].iloc[0]
                            if isinstance(hinge_jamb, dict):
                                hinge_jamb = hinge_jamb.get('value', '')
                    
                    if frame_header_id:
                        frame_header_attr = task_attrs[task_attrs['task_type_attribute_id'] == frame_header_id]
                        if not frame_header_attr.empty:
                            frame_header = frame_header_attr['text_value'].iloc[0]
                            if isinstance(frame_header, dict):
                                frame_header = frame_header.get('value', '')
                
                # Get latest updated_at or created_at
                latest_date = None
                for attr in [strike_jamb_attr, hinge_jamb_attr, frame_header_attr]:
                    if not attr.empty:
                        # Try updated_at first
                        if 'updated_at' in attr and attr['updated_at'].iloc[0]:
                            date = pd.to_datetime(attr['updated_at'].iloc[0])
                            # Convert to timezone-unaware datetime
                            if date.tz is not None:
                                date = date.tz_localize(None)
                            if latest_date is None or date > latest_date:
                                latest_date = date
                        # Fall back to created_at if updated_at is not available
                        elif 'created_at' in attr and attr['created_at'].iloc[0]:
                            date = pd.to_datetime(attr['created_at'].iloc[0])
                            # Convert to timezone-unaware datetime
                            if date.tz is not None:
                                date = date.tz_localize(None)
                            if latest_date is None or date > latest_date:
                                latest_date = date
                
                result_data.append({
                    'Opening': task['name'],
                    'Strike Jamb': strike_jamb,
                    'Hinge Jamb': hinge_jamb,
                    'Frame Header': frame_header,
                    'Date Submitted': latest_date if latest_date else ''
                })
            
            # Create final DataFrame
            result_df = pd.DataFrame(result_data)
            
            # Export to Excel
            self.export_to_excel(output_filename, result_df)
            
        except Exception as e:
            raise ValueError(f"Failed to generate FC task report: {str(e)}") 