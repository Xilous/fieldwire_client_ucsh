import json
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path
from src.models import ComparisonSummary, OpeningChange, HardwareChange
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

class Exporter:
    """Exports comparison results to JSON and Excel formats."""
    
    @staticmethod
    def export_to_json(
        summary: ComparisonSummary,
        old_file: str,
        new_file: str,
        output_file: str
    ):
        """Export comparison results to JSON file."""
        
        # Create export data structure
        export_data = {
            "metadata": {
                "oldVersion": old_file,
                "newVersion": new_file,
                "comparisonDate": datetime.now().isoformat(),
                "summary": {
                    "totalChangedOpenings": summary.total_changed_openings,
                    "openingsWithDoorInfoChanges": summary.openings_with_door_info_changes,
                    "openingsWithHardwareChanges": summary.openings_with_hardware_changes
                }
            },
            "changes": {
                "openings": [
                    Exporter._format_opening_change(change)
                    for change in summary.changes
                ]
            }
        }
        
        # Write to JSON file
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2)
            
        # Create Excel file
        excel_path = Path(output_file).with_suffix('.xlsx')
        Exporter.export_to_excel(summary, old_file, new_file, str(excel_path))
    
    @staticmethod
    def export_to_excel(
        summary: ComparisonSummary,
        old_file: str,
        new_file: str,
        output_file: str
    ):
        """Export comparison results to Excel file with multiple sheets."""
        wb = Workbook()
        
        # Create styles
        header_style = Font(bold=True)
        added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        Exporter._create_summary_sheet(ws_summary, summary, old_file, new_file)
        
        # Door Info Changes Sheet
        ws_door = wb.create_sheet("Door Info Changes")
        Exporter._create_door_info_sheet(ws_door, summary.changes, header_style, modified_fill)
        
        # Hardware Changes Sheet
        ws_hardware = wb.create_sheet("Hardware Changes")
        Exporter._create_hardware_sheet(
            ws_hardware, 
            summary.changes,
            header_style,
            added_fill,
            deleted_fill,
            modified_fill
        )
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_door, ws_hardware]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_file)
    
    @staticmethod
    def _create_summary_sheet(ws, summary: ComparisonSummary, old_file: str, new_file: str):
        """Create the summary sheet."""
        ws.append(["Door Hardware Schedule Comparison Summary"])
        ws.append([])
        ws.append(["Old Version", Path(old_file).name])
        ws.append(["New Version", Path(new_file).name])
        ws.append(["Comparison Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        ws.append(["Statistics"])
        ws.append(["Total Changed Openings", summary.total_changed_openings])
        ws.append(["Openings with Door Info Changes", summary.openings_with_door_info_changes])
        ws.append(["Openings with Hardware Changes", summary.openings_with_hardware_changes])
    
    @staticmethod
    def _create_door_info_sheet(ws, changes: List[OpeningChange], header_style, modified_fill):
        """Create the door information changes sheet."""
        # Headers
        headers = ["Opening Number", "Field", "Old Value", "New Value"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_style
        
        # Data
        for change in changes:
            if change.door_info_changes:
                for info_change in change.door_info_changes:
                    row = [
                        change.number,
                        info_change.field,
                        info_change.old_value,
                        info_change.new_value
                    ]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.fill = modified_fill
    
    @staticmethod
    def _create_hardware_sheet(ws, changes: List[OpeningChange], header_style, added_fill, deleted_fill, modified_fill):
        """Create the hardware changes sheet."""
        current_row = 1
        
        for change in changes:
            if not change.hardware_changes:
                continue
                
            # Opening header
            ws.append([f"Opening {change.number}"])
            ws.cell(current_row, 1).font = header_style
            current_row += 1
            
            # Added Hardware
            added = [c for c in change.hardware_changes if c.type == 'added']
            if added:
                current_row = Exporter._add_hardware_section(
                    ws, "Added Hardware", added, current_row, header_style, added_fill
                )
            
            # Deleted Hardware
            deleted = [c for c in change.hardware_changes if c.type == 'deleted']
            if deleted:
                current_row = Exporter._add_hardware_section(
                    ws, "Deleted Hardware", deleted, current_row, header_style, deleted_fill
                )
            
            # Modified Hardware
            modified = [c for c in change.hardware_changes if c.type == 'modified']
            if modified:
                ws.append(["Modified Hardware"])
                ws.cell(current_row, 1).font = header_style
                current_row += 1
                
                headers = ["Short Code", "Product Code", "Sub Category", "Field", "Old Value", "New Value"]
                ws.append(headers)
                for cell in ws[current_row]:
                    cell.font = header_style
                current_row += 1
                
                for change in modified:
                    for field, values in change.modifications.items():
                        row = [
                            change.item.short_code,
                            change.item.product_code,
                            change.item.sub_category,
                            field,
                            values['old'],
                            values['new']
                        ]
                        ws.append(row)
                        for cell in ws[current_row]:
                            cell.fill = modified_fill
                        current_row += 1
            
            # Add spacing between openings
            ws.append([])
            current_row += 1
    
    @staticmethod
    def _add_hardware_section(ws, title: str, items: List[HardwareChange], start_row: int, header_style, fill_style) -> int:
        """Add a section of hardware items to the sheet."""
        ws.append([title])
        ws.cell(start_row, 1).font = header_style
        current_row = start_row + 1
        
        headers = ["Short Code", "Product Code", "Sub Category", "Quantity", "Handing", "Finish"]
        ws.append(headers)
        for cell in ws[current_row]:
            cell.font = header_style
        current_row += 1
        
        for change in items:
            row = [
                change.item.short_code,
                change.item.product_code,
                change.item.sub_category,
                change.item.quantity_active,
                change.item.handing,
                change.item.finish_ansi
            ]
            ws.append(row)
            for cell in ws[current_row]:
                cell.fill = fill_style
            current_row += 1
        
        ws.append([])
        return current_row + 1
    
    @staticmethod
    def _format_opening_change(change) -> Dict[str, Any]:
        """Format opening change for JSON export."""
        return {
            "number": change.number,
            "doorInfo": {
                "modified": {
                    info_change.field: {
                        "old": info_change.old_value,
                        "new": info_change.new_value
                    }
                    for info_change in change.door_info_changes
                }
            } if change.door_info_changes else {},
            "hardware": {
                "added": [
                    Exporter._format_hardware_item(c.item)
                    for c in change.hardware_changes
                    if c.type == 'added'
                ],
                "deleted": [
                    Exporter._format_hardware_item(c.item)
                    for c in change.hardware_changes
                    if c.type == 'deleted'
                ],
                "modified": [
                    {
                        "identifier": {
                            "shortCode": c.item.short_code,
                            "productCode": c.item.product_code,
                            "subCategory": c.item.sub_category
                        },
                        "changes": c.modifications
                    }
                    for c in change.hardware_changes
                    if c.type == 'modified'
                ]
            }
        }
    
    @staticmethod
    def _format_hardware_item(item) -> Dict[str, Any]:
        """Format hardware item for JSON export."""
        return {
            "shortCode": item.short_code,
            "productCode": item.product_code,
            "subCategory": item.sub_category,
            "quantityActive": item.quantity_active,
            "handing": item.handing,
            "finishANSI": item.finish_ansi
        }